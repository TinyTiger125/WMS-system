import base64
import io

import openpyxl
import xlsxwriter

from odoo import fields, models
from odoo.exceptions import ValidationError


class CustomProductImportWizard(models.TransientModel):
    _name = "custom.product.import.wizard"
    _description = "Product Import Wizard"

    upload_file = fields.Binary(string="导入文件", required=True)
    upload_filename = fields.Char(string="文件名")
    result_note = fields.Text(string="导入结果", readonly=True)
    template_file = fields.Binary(string="模板文件", readonly=True)
    template_filename = fields.Char(string="模板文件名", readonly=True)

    @staticmethod
    def _template_headers():
        return [
            "条码",
            "中文名",
            "日文名",
            "品牌名称",
            "生产厂家",
            "规格",
            "成分",
            "产地",
            "申报价格",
            "销售价格",
            "净重(kg)",
            "毛重(kg)",
            "箱规",
            "菜鸟商品ID",
            "日本海关分类",
            "中国海关分类",
            "单位1",
            "单位2",
            "备注",
            "HS Code",
            "中文报关名",
            "日文报关名",
        ]

    def action_download_template(self):
        self.ensure_one()
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("商品导入模板")
        header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        cell_fmt = workbook.add_format({"border": 1})

        headers = self._template_headers()
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_fmt)
            worksheet.set_column(col, col, 18)

        sample = [
            "4901417163059",
            "测试商品300ml",
            "テスト商品300ml",
            "测试品牌",
            "测试厂家",
            "300ml/瓶",
            "成分A, 成分B",
            "日本",
            89.0,
            199.0,
            0.3,
            0.35,
            "36瓶/箱",
            "CAINIAO-001",
            "护肤品",
            "化妆品",
            "瓶",
            "箱",
            "导入示例",
            "33049900",
            "中文报关名示例",
            "日文报关名サンプル",
        ]
        for col, value in enumerate(sample):
            worksheet.write(1, col, value, cell_fmt)

        workbook.close()
        self.template_file = base64.b64encode(output.getvalue())
        self.template_filename = "商品导入模板.xlsx"
        return self.action_download_generated_template()

    def action_download_generated_template(self):
        self.ensure_one()
        if not self.template_file:
            raise ValidationError("请先生成模板。")
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content?model={self._name}&id={self.id}&field=template_file&filename_field=template_filename&download=true",
            "target": "self",
        }

    @staticmethod
    def _clean_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _to_float(value):
        if value in (None, ""):
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _find_country(self, country_value):
        value = self._clean_str(country_value)
        if not value:
            return False
        alias_code = {
            "日本": "JP",
            "日本国": "JP",
            "japan": "JP",
            "中国": "CN",
            "中国大陆": "CN",
            "china": "CN",
            "韩国": "KR",
            "korea": "KR",
            "美国": "US",
            "usa": "US",
            "united states": "US",
        }.get(value.lower(), value.upper())

        country = self.env["res.country"].search([("code", "=", alias_code)], limit=1)
        if country:
            return country
        country = self.env["res.country"].with_context(lang="zh_CN").search([("name", "ilike", value)], limit=1)
        if country:
            return country
        return self.env["res.country"].search([("name", "ilike", value)], limit=1)

    def _find_customs_category(self, category_value, category_type):
        value = self._clean_str(category_value)
        if not value:
            return False
        category = self.env["custom.customs.category"].search(
            [("category_type", "=", category_type), ("name", "=", value)],
            limit=1,
        )
        if category:
            return category
        return self.env["custom.customs.category"].search(
            [("category_type", "=", category_type), ("code", "=", value)],
            limit=1,
        )

    def action_import_products(self):
        self.ensure_one()
        if not self.upload_file:
            raise ValidationError("请先上传导入文件。")

        content = base64.b64decode(self.upload_file)
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("导入文件为空。")

        headers = [self._clean_str(col) for col in rows[0]]
        required_headers = {"条码", "中文名", "规格", "产地"}
        if not required_headers.issubset(set(headers)):
            missing = required_headers.difference(set(headers))
            raise ValidationError(f"导入模板缺少必需列：{', '.join(sorted(missing))}")

        header_index = {header: idx for idx, header in enumerate(headers) if header}
        product_model = self.env["product.template"].with_context(active_test=False)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            try:
                row_data = {}
                for header, idx in header_index.items():
                    row_data[header] = row[idx] if idx < len(row) else None

                barcode = self._clean_str(row_data.get("条码"))
                product_name = self._clean_str(row_data.get("中文名"))
                if not barcode or not product_name:
                    skipped_count += 1
                    continue

                country = self._find_country(row_data.get("产地"))
                if not country:
                    raise ValidationError("未匹配到有效的产地。")

                jp_category = self._find_customs_category(row_data.get("日本海关分类"), "jp")
                cn_category = self._find_customs_category(row_data.get("中国海关分类"), "cn")

                vals = {
                    "type": "consu",
                    "is_storable": True,
                    "sale_ok": True,
                    "purchase_ok": True,
                    "barcode": barcode,
                    "name": product_name,
                    "name_jp": self._clean_str(row_data.get("日文名")),
                    "brand_name": self._clean_str(row_data.get("品牌名称")),
                    "manufacturer_name": self._clean_str(row_data.get("生产厂家")),
                    "spec_text": self._clean_str(row_data.get("规格")),
                    "ingredient_text": self._clean_str(row_data.get("成分")),
                    "origin_country_id": country.id,
                    "declared_value": self._to_float(row_data.get("申报价格")),
                    "list_price": self._to_float(row_data.get("销售价格")),
                    "net_weight_kg": self._to_float(row_data.get("净重(kg)")),
                    "gross_weight_kg": self._to_float(row_data.get("毛重(kg)")),
                    "carton_spec": self._clean_str(row_data.get("箱规")),
                    "cainiao_product_id": self._clean_str(row_data.get("菜鸟商品ID")),
                    "unit_1": self._clean_str(row_data.get("单位1")),
                    "unit_2": self._clean_str(row_data.get("单位2")),
                    "remark": self._clean_str(row_data.get("备注")),
                    "customs_hs_code": self._clean_str(row_data.get("HS Code")),
                    "customs_name_cn": self._clean_str(row_data.get("中文报关名")),
                    "customs_name_jp": self._clean_str(row_data.get("日文报关名")),
                    "jp_customs_category_id": jp_category.id if jp_category else False,
                    "cn_customs_category_id": cn_category.id if cn_category else False,
                }

                product = product_model.search([("barcode", "=", barcode)], limit=1)
                if product:
                    product.write(vals)
                    updated_count += 1
                else:
                    product_model.create(vals)
                    created_count += 1
            except Exception as exc:
                errors.append(f"第 {row_number} 行：{exc}")

        result_lines = [
            f"新建：{created_count}",
            f"更新：{updated_count}",
            f"跳过：{skipped_count}",
            f"错误：{len(errors)}",
        ]
        if errors:
            result_lines.append("错误明细：")
            result_lines.extend(errors[:30])

        self.result_note = "\n".join(result_lines)
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
